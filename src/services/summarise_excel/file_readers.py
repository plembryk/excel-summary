import abc

from typing import Generator

from openpyxl import load_workbook

from services.summarise_excel.exceptions import CannotReadFileError


class BaseFileReader(abc.ABC):
    @abc.abstractmethod
    def iter_rows(self, min_row: int = 0) -> Generator: ...


class ExcelFileReader(BaseFileReader):
    def __init__(self, file: str):
        self.file = file

    def iter_rows(self, min_row: int = 0) -> Generator:
        workbook = load_workbook(self.file, read_only=True)
        try:
            worksheet = workbook.active

            if worksheet is not None:
                for row in worksheet.iter_rows(values_only=True, min_row=min_row):
                    yield row
        except Exception:
            raise CannotReadFileError
        finally:
            workbook.close()
